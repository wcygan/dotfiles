#!/usr/bin/env -S uv run --script
# /// script
# requires-python = ">=3.12"
# dependencies = [
#   "defusedxml>=0.7,<1",
#   "duckdb>=1.0,<2",
#   "openpyxl>=3.1,<4",
#   "pillow>=10,<13",
#   "pypdf>=5,<7",
#   "pyyaml>=6,<7",
#   "rich>=13,<15",
# ]
# ///
"""Gold-standard file intake script for common automation formats.

Run:
    uv run --script inspect_files.py --demo
    uv run --script inspect_files.py data.json report.xlsx image.png
"""

from __future__ import annotations

import argparse
import csv
import json
import tempfile
import tomllib
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import duckdb
import yaml
from defusedxml import ElementTree
from openpyxl import Workbook, load_workbook
from PIL import Image
from pypdf import PdfReader, PdfWriter
from rich.console import Console
from rich.table import Table

console = Console()
error_console = Console(stderr=True)


@dataclass(frozen=True)
class FileSummary:
    path: Path
    kind: str
    details: dict[str, Any]


def summarize_json(path: Path) -> dict[str, Any]:
    data = json.loads(path.read_text())
    if isinstance(data, list):
        return {"type": "array", "items": len(data)}
    if isinstance(data, dict):
        return {"type": "object", "keys": len(data), "sample_keys": list(data)[:5]}
    return {"type": type(data).__name__}


def summarize_yaml(path: Path) -> dict[str, Any]:
    data = yaml.safe_load(path.read_text())
    if isinstance(data, dict):
        return {"type": "object", "keys": len(data), "sample_keys": list(data)[:5]}
    return {"type": type(data).__name__}


def summarize_toml(path: Path) -> dict[str, Any]:
    data = tomllib.loads(path.read_text())
    return {"type": "object", "keys": len(data), "sample_keys": list(data)[:5]}


def summarize_table(path: Path) -> dict[str, Any]:
    with duckdb.connect(":memory:") as connection:
        suffix = path.suffix.lower()
        if suffix == ".csv":
            relation = connection.read_csv(str(path), header=True)
        elif suffix in {".json", ".jsonl", ".ndjson"}:
            relation = connection.read_json(str(path))
        elif suffix == ".parquet":
            relation = connection.read_parquet(str(path))
        else:
            raise ValueError(f"unsupported tabular file: {path}")
        return {
            "columns": len(relation.columns),
            "rows": relation.count("*").fetchone()[0],
            "column_names": relation.columns[:5],
        }


def summarize_excel(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    return {
        "sheets": workbook.sheetnames,
        "active_rows": workbook.active.max_row,
        "active_columns": workbook.active.max_column,
    }


def summarize_xml(path: Path) -> dict[str, Any]:
    root = ElementTree.parse(path).getroot()
    return {"root": root.tag, "children": len(list(root))}


def summarize_pdf(path: Path) -> dict[str, Any]:
    reader = PdfReader(path)
    return {"pages": len(reader.pages), "encrypted": reader.is_encrypted}


def summarize_image(path: Path) -> dict[str, Any]:
    with Image.open(path) as image:
        return {
            "format": image.format,
            "width": image.width,
            "height": image.height,
            "mode": image.mode,
        }


def summarize_file(path: Path) -> FileSummary:
    suffix = path.suffix.lower()
    if suffix in {".json", ".jsonl", ".ndjson"}:
        details = summarize_table(path) if suffix != ".json" else summarize_json(path)
        return FileSummary(path, "json", details)
    if suffix in {".yaml", ".yml"}:
        return FileSummary(path, "yaml", summarize_yaml(path))
    if suffix == ".toml":
        return FileSummary(path, "toml", summarize_toml(path))
    if suffix == ".csv":
        return FileSummary(path, "csv", summarize_table(path))
    if suffix == ".parquet":
        return FileSummary(path, "parquet", summarize_table(path))
    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return FileSummary(path, "excel", summarize_excel(path))
    if suffix == ".xml":
        return FileSummary(path, "xml", summarize_xml(path))
    if suffix == ".pdf":
        return FileSummary(path, "pdf", summarize_pdf(path))
    if suffix in {".png", ".jpg", ".jpeg", ".webp", ".tif", ".tiff"}:
        return FileSummary(path, "image", summarize_image(path))
    raise ValueError(f"unsupported file type: {path}")


def create_demo_files(directory: Path) -> list[Path]:
    json_path = directory / "sample.json"
    json_path.write_text(json.dumps({"name": "demo", "items": [1, 2, 3]}))

    yaml_path = directory / "sample.yaml"
    yaml_path.write_text("name: demo\nitems:\n  - alpha\n  - beta\n")

    toml_path = directory / "sample.toml"
    toml_path.write_text("[tool.demo]\nname = 'demo'\n")

    csv_path = directory / "sample.csv"
    with csv_path.open("w", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["name", "count"])
        writer.writerow(["alpha", 1])
        writer.writerow(["beta", 2])

    workbook_path = directory / "sample.xlsx"
    workbook = Workbook()
    workbook.active.append(["name", "count"])
    workbook.active.append(["alpha", 1])
    workbook.save(workbook_path)

    xml_path = directory / "sample.xml"
    xml_path.write_text("<root><item id='1'/><item id='2'/></root>")

    pdf_path = directory / "sample.pdf"
    writer = PdfWriter()
    writer.add_blank_page(width=72, height=72)
    with pdf_path.open("wb") as handle:
        writer.write(handle)

    image_path = directory / "sample.png"
    Image.new("RGB", (32, 16), color="navy").save(image_path)

    return [
        json_path,
        yaml_path,
        toml_path,
        csv_path,
        workbook_path,
        xml_path,
        pdf_path,
        image_path,
    ]


def render(summaries: list[FileSummary]) -> None:
    table = Table(title="File Inspection")
    table.add_column("Path")
    table.add_column("Kind")
    table.add_column("Details")
    for summary in summaries:
        table.add_row(
            str(summary.path), summary.kind, json.dumps(summary.details, sort_keys=True)
        )
    console.print(table)


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Inspect common file formats safely.")
    parser.add_argument("paths", nargs="*", type=Path)
    parser.add_argument(
        "--demo", action="store_true", help="Inspect generated sample files."
    )
    return parser.parse_args(argv)


def run(args: argparse.Namespace) -> int:
    if args.demo:
        with tempfile.TemporaryDirectory() as tmp:
            render([summarize_file(path) for path in create_demo_files(Path(tmp))])
        return 0

    if not args.paths:
        error_console.print("[red]missing paths[/red] (or use --demo)")
        return 2

    missing = [path for path in args.paths if not path.exists()]
    if missing:
        for path in missing:
            error_console.print(f"[red]missing file[/red] {path}")
        return 2

    render([summarize_file(path) for path in args.paths])
    return 0


def main(argv: list[str] | None = None) -> int:
    try:
        return run(parse_args(argv))
    except (OSError, ValueError, yaml.YAMLError, duckdb.Error) as exc:
        error_console.print(f"[red]inspection failed[/red] {exc}")
        return 1
    except KeyboardInterrupt:
        error_console.print("[red]interrupted[/red]")
        return 130


if __name__ == "__main__":
    raise SystemExit(main())
